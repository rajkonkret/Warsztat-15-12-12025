from openpyxl import Workbook

# pip install openpyxl

wb = Workbook()
ws = wb.active
ws.title = "Dane"
ws['A1'] = "Item"
ws['B1'] = "Price"

ws.append(["Pen", 10])
ws.append(["Notebook", 25])

wb.save("data.xlsx")
